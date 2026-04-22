import PyPDF2
import openpyxl
from pathlib import Path
from typing import List, Dict, Any
import re

class DocumentParser:
    """Parse PDF and Excel documents into chunks"""
    
    @staticmethod
    def parse_pdf(file_path: str) -> List[str]:
        """Extract text from PDF"""
        texts = []
        try:
            with open(file_path, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                for page_num, page in enumerate(pdf_reader.pages):
                    text = page.extract_text()
                    if text.strip():
                        # Add page marker for context
                        texts.append(f"[Page {page_num + 1}]\n{text}")
        except Exception as e:
            print(f"Error parsing PDF {file_path}: {e}")
        return texts
    
    @staticmethod
    def parse_excel(file_path: str) -> List[str]:
        """Extract data from Excel as structured text"""
        texts = []
        try:
            wb = openpyxl.load_workbook(file_path)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                texts.append(f"[Sheet: {sheet_name}]\n")
                
                # Get headers
                headers = []
                for cell in ws[1]:
                    headers.append(str(cell.value))
                
                if headers:
                    texts.append("Headers: " + " | ".join(headers) + "\n")
                
                # Get data rows
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                    row_data = []
                    for cell in row:
                        row_data.append(str(cell.value))
                    texts.append(f"Row {row_idx}: " + " | ".join(row_data))
        except Exception as e:
            print(f"Error parsing Excel {file_path}: {e}")
        return texts
    
    @staticmethod
    def chunk_text(texts: List[str], chunk_size: int = 512, overlap: int = 50) -> List[str]:
        """Split texts into overlapping chunks"""
        chunks = []
        full_text = "\n".join(texts)
        
        # Split by sentences first (rough approximation based on words)
        words = full_text.split()
        
        for i in range(0, len(words), chunk_size - overlap):
            chunk = " ".join(words[i:i + chunk_size])
            if chunk.strip():
                chunks.append(chunk)
        
        return chunks
    
    @staticmethod
    def load_and_process_documents(data_dir: str = ".") -> List[Dict[str, Any]]:
        """Load all documents from directory and return processed chunks"""
        documents = []
        data_path = Path(data_dir)
        
        # Process PDF files
        for pdf_file in data_path.glob("*.pdf"):
            print(f"Processing PDF: {pdf_file.name}")
            texts = DocumentParser.parse_pdf(str(pdf_file))
            chunks = DocumentParser.chunk_text(texts)
            for chunk in chunks:
                documents.append({
                    "content": chunk,
                    "source": pdf_file.name,
                    "type": "pdf"
                })
        
        # Process Excel files
        for excel_file in data_path.glob("*.xlsx"):
            print(f"Processing Excel: {excel_file.name}")
            texts = DocumentParser.parse_excel(str(excel_file))
            chunks = DocumentParser.chunk_text(texts)
            for chunk in chunks:
                documents.append({
                    "content": chunk,
                    "source": excel_file.name,
                    "type": "xlsx"
                })
        
        print(f"Loaded {len(documents)} document chunks")
        return documents
